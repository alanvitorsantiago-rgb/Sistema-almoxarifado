# /app.py

import os
from sqlalchemy import or_, func
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response, session
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
# from flask_socketio import SocketIO  # Temporariamente desabilitado
from flask_bcrypt import Bcrypt
from models import db, User, ItemEstoque, Movimentacao, EstoqueDetalhe, ConsumivelEstoque, MovimentacaoConsumivel
from functools import wraps
from datetime import datetime, date, timedelta
import pandas as pd
import numpy as np
# Importa o modelo de suavização exponencial para previsão
from statsmodels.tsa.api import ExponentialSmoothing
import io
from flask_socketio import SocketIO
import unicodedata

# --- INICIALIZAÇÃO E CONFIGURAÇÃO DA APLICAÇÃO ---
app = Flask(__name__)
basedir = os.path.abspath(os.path.dirname(__file__))

# Carregar variáveis de ambiente
from dotenv import load_dotenv
load_dotenv('.env.supabase')  # Carrega configurações do arquivo .env.supabase

def configure_app(app_instance):
    """Configura a aplicação Flask."""
    # Chave secreta para sessões e mensagens flash
    app_instance.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'sua-chave-secreta-super-segura')
    
    # Configuração do banco de dados PostgreSQL (Supabase)
    database_url = os.getenv('DATABASE_URL')
    if database_url and database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    
    app_instance.config['SQLALCHEMY_DATABASE_URI'] = database_url or 'sqlite:///' + os.path.join(basedir, 'database.db')
    app_instance.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

    # Inicializa as extensões com a aplicação
    db.init_app(app_instance)
    bcrypt.init_app(app_instance)
    socketio.init_app(app_instance)
    login_manager.init_app(app_instance)

# Inicializa as extensões sem a aplicação para serem configuradas depois
bcrypt = Bcrypt()
login_manager = LoginManager()
socketio = SocketIO()
login_manager.login_view = 'login'  # Rota para redirecionar usuários não logados
login_manager.login_message = "Por favor, faça login para acessar esta página."
login_manager.login_message_category = 'info'

configure_app(app)

# --- FUNÇÃO PARA CRIAR O BANCO DE DADOS ---
def criar_banco_de_dados():
    """Cria as tabelas do banco de dados se não existirem."""
    db.create_all()
    print("✅ Tabelas do banco de dados verificadas/criadas com sucesso!")
    
    # Garante que o usuário 'admin' exista
    admin_user = User.query.filter_by(username='admin').first()
    if not admin_user:
        print("🔧 Usuário 'admin' não encontrado. Criando agora...")
        hashed_password = bcrypt.generate_password_hash('admin').decode('utf-8')
        new_admin = User(username='admin', password_hash=hashed_password, role='admin')
        db.session.add(new_admin)
        db.session.commit()
        print("✅ Usuário 'admin' criado com a senha 'admin'.")

# Cria as tabelas do banco de dados na inicialização
with app.app_context():
    criar_banco_de_dados()

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# --- FUNÇÕES AUXILIARES E CONTEXT PROCESSORS ---

def calculate_validity_status(validade_date):
    """Calcula o status de validade e retorna uma tupla (texto, classe_css)."""
    if not validade_date:
        return ("Sem validade", "text-muted")

    today = datetime.today().date()
    delta = (validade_date - today).days

    if delta < 0:
        return (f"Vencido há {-delta} dia(s)", "text-danger fw-bold")
    elif delta == 0:
        return ("Vence hoje!", "text-danger fw-bold")
    elif delta <= 30:
        return (f"Vence em {delta} dia(s)", "text-warning fw-bold")
    elif delta <= 90:
        return (f"Vence em {delta} dia(s)", "text-info")
    else:
        return (f"Vence em {delta} dia(s)", "text-success")

@app.context_processor
def utility_processor():
    """Disponibiliza a função para todos os templates."""
    return dict(calculate_status=calculate_validity_status)

# --- DECORADOR PARA ROTAS DE ADMIN ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Acesso negado. Esta área é restrita a administradores.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# --- DECORADOR PARA ROTAS DE ADMIN ONLY (Nega acesso a usuários padrão) ---
def admin_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Acesso negado. Esta ação é permitida apenas para administradores.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


# --- ROTAS DA APLICAÇÃO ---

@app.route('/relatorio/movimentacoes')
@admin_only
@login_required
def relatorio_movimentacoes():
    """Exibe o relatório de movimentações com filtros de data."""
    data_inicio_str = request.args.get('data_inicio')
    data_fim_str = request.args.get('data_fim')
    search_query = request.args.get('q', '') # Pega o parâmetro de busca 'q'

    # Inicia a query com um join para permitir a busca nos dados do item
    query = Movimentacao.query.join(ItemEstoque, Movimentacao.item_id == ItemEstoque.id)

    if data_inicio_str:
        data_inicio = datetime.strptime(data_inicio_str, '%Y-%m-%d')
        query = query.filter(Movimentacao.data_movimentacao >= data_inicio)
    if data_fim_str:
        # Adiciona o final do dia para incluir todos os registros da data final
        data_fim = datetime.strptime(f'{data_fim_str} 23:59:59', '%Y-%m-%d %H:%M:%S')
        query = query.filter(Movimentacao.data_movimentacao <= data_fim)
    
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(
            or_(
                ItemEstoque.codigo.ilike(search_term),
                ItemEstoque.descricao.ilike(search_term),
                Movimentacao.lote.ilike(search_term),
                Movimentacao.usuario.ilike(search_term),
                Movimentacao.observacao.ilike(search_term)
            )
        )

    movimentacoes = query.order_by(Movimentacao.data_movimentacao.desc()).all()

    # --- Lógica para o Gráfico de Pizza ---
    # Conta o número de ocorrências de cada tipo de movimentação na lista filtrada
    entradas_count = sum(1 for mov in movimentacoes if 'ENTRADA' in mov.tipo and 'AJUSTE' not in mov.tipo)
    saidas_count = sum(1 for mov in movimentacoes if 'SAIDA' in mov.tipo and 'AJUSTE' not in mov.tipo)
    ajustes_count = sum(1 for mov in movimentacoes if 'AJUSTE' in mov.tipo)

    pie_chart_data = {
        'labels': ['Entradas', 'Saídas', 'Ajustes'],
        'counts': [entradas_count, saidas_count, ajustes_count]
    }

    return render_template('relatorio_movimentacoes.html', 
                           movimentacoes=movimentacoes, 
                           title="Relatório de Movimentações", 
                           data_inicio=data_inicio_str, data_fim=data_fim_str, 
                           search_query=search_query,
                           pie_chart_data=pie_chart_data)

@app.route('/dashboard')
@login_required
def dashboard():
    """Renderiza o dashboard com dados reais do banco de dados."""
    hoje = date.today()

    # --- KPIs (Key Performance Indicators) ---
    total_items_distintos = ItemEstoque.query.count()
    total_unidades = db.session.query(func.sum(ItemEstoque.qtd_estoque)).scalar() or 0
    itens_zerados = ItemEstoque.query.filter(ItemEstoque.qtd_estoque == 0).count()
    
    # Lotes críticos (vencem nos próximos 30 dias ou já venceram)
    data_limite_criticos = hoje + timedelta(days=30)
    critical_lotes = EstoqueDetalhe.query.filter(
        EstoqueDetalhe.validade.isnot(None),
        EstoqueDetalhe.validade <= data_limite_criticos,
        EstoqueDetalhe.quantidade > 0
    ).all()

    # --- Dados para o Gráfico de Movimentações (Últimos 15 dias) ---
    labels_mov = [(hoje - timedelta(days=i)).strftime('%d/%m') for i in range(14, -1, -1)]
    entradas = []
    saidas = []
    for i in range(14, -1, -1):
        dia = hoje - timedelta(days=i)
        entradas.append(db.session.query(func.sum(Movimentacao.quantidade)).filter(Movimentacao.tipo=='ENTRADA', func.date(Movimentacao.data_movimentacao)==dia).scalar() or 0)
        saidas.append(db.session.query(func.sum(Movimentacao.quantidade)).filter(Movimentacao.tipo=='SAIDA', func.date(Movimentacao.data_movimentacao)==dia).scalar() or 0)

    movimentacoes_chart_data = {
        'labels': labels_mov, 'entradas': entradas, 'saidas': saidas
    }

    # --- Dados para o Gráfico de Tipos (Pizza) ---
    tipos_data = db.session.query(ItemEstoque.tipo, func.count(ItemEstoque.id)).group_by(ItemEstoque.tipo).order_by(func.count(ItemEstoque.id).desc()).all()
    tipos_chart_data = {
        'labels': [tipo if tipo else 'Não categorizado' for tipo, count in tipos_data],
        'counts': [count for tipo, count in tipos_data]
    }

    # --- Listas Top 5 ---
    # Filtrar apenas itens com descrição válida (não vazia, não apenas "-")
    top_stocked_items = ItemEstoque.query.filter(
        ItemEstoque.descricao != "",
        ItemEstoque.descricao != "-",
        ItemEstoque.descricao != "="
    ).order_by(ItemEstoque.qtd_estoque.desc()).limit(5).all()
    
    low_stocked_items = ItemEstoque.query.filter(
        ItemEstoque.qtd_estoque > 0, 
        ItemEstoque.qtd_estoque <= ItemEstoque.estoque_minimo,
        ItemEstoque.descricao != "",
        ItemEstoque.descricao != "-",
        ItemEstoque.descricao != "="
    ).order_by(ItemEstoque.qtd_estoque.asc()).limit(5).all()

    # --- Atividade Recente ---
    recent_movimentacoes = Movimentacao.query.order_by(Movimentacao.data_movimentacao.desc()).limit(5).all()

    # --- Dados para Alertas de Validade ---
    data_limite_proximos = hoje + timedelta(days=40)

    # Lotes que vencem exatamente hoje e têm quantidade
    lotes_vencendo_hoje = EstoqueDetalhe.query.filter(
        EstoqueDetalhe.validade == hoje,
        EstoqueDetalhe.quantidade > 0
    ).order_by(EstoqueDetalhe.item_estoque_id).all()

    # Lotes que vencem nos próximos 40 dias (excluindo hoje) e têm quantidade
    lotes_proximo_vencimento = EstoqueDetalhe.query.filter(
        EstoqueDetalhe.validade > hoje,
        EstoqueDetalhe.validade <= data_limite_proximos,
        EstoqueDetalhe.quantidade > 0
    ).order_by(EstoqueDetalhe.validade.asc()).all()

    # --- DADOS DE CONSUMÍVEIS ---
    total_consumiveis = ConsumivelEstoque.query.count()
    consumiveis_zerados = ConsumivelEstoque.query.filter(ConsumivelEstoque.quantidade_atual == 0).count()
    consumiveis_baixo_estoque = ConsumivelEstoque.query.filter(
        ConsumivelEstoque.quantidade_atual > 0,
        ConsumivelEstoque.quantidade_atual <= ConsumivelEstoque.estoque_minimo
    ).count()
    
    # Top 5 consumíveis com menor quantidade (alerta visual)
    low_consumiveis = ConsumivelEstoque.query.order_by(ConsumivelEstoque.quantidade_atual.asc()).limit(5).all()
    
    # Últimas 5 movimentações de consumíveis
    recent_consumivel_moves = MovimentacaoConsumivel.query.order_by(MovimentacaoConsumivel.data_movimentacao.desc()).limit(5).all()

    return render_template('dashboard.html', 
                           total_items_distintos=total_items_distintos,
                           total_unidades=total_unidades,
                           itens_zerados=itens_zerados,
                           critical_lotes=critical_lotes,
                           movimentacoes_chart_data=movimentacoes_chart_data,
                           tipos_chart_data=tipos_chart_data,
                           top_stocked_items=top_stocked_items,
                           low_stocked_items=low_stocked_items,
                           recent_movimentacoes=recent_movimentacoes,
                           lotes_vencendo_hoje=lotes_vencendo_hoje,
                           lotes_proximo_vencimento=lotes_proximo_vencimento,
                           today_date=hoje,
                           total_consumiveis=total_consumiveis,
                           consumiveis_zerados=consumiveis_zerados,
                           consumiveis_baixo_estoque=consumiveis_baixo_estoque,
                           low_consumiveis=low_consumiveis,
                           recent_consumivel_moves=recent_consumivel_moves)

@app.route('/api/kpis')
@login_required
def api_kpis():
    """Retorna os dados dos KPIs principais em formato JSON."""
    hoje = date.today()
    data_limite_criticos = hoje + timedelta(days=30)
    
    kpis = {
        'total_items_distintos': ItemEstoque.query.count(),
        'total_unidades': db.session.query(func.sum(ItemEstoque.qtd_estoque)).scalar() or 0,
        'itens_zerados': ItemEstoque.query.filter(ItemEstoque.qtd_estoque == 0).count(),
        'critical_lotes': EstoqueDetalhe.query.filter(
            EstoqueDetalhe.validade.isnot(None),
            EstoqueDetalhe.validade <= data_limite_criticos,
            EstoqueDetalhe.quantidade > 0
        ).count()
    }
    return jsonify(kpis)

@app.route('/api/items/search')
@login_required
def api_items_search():
    """API para buscar itens por código ou descrição para autocompletar."""
    search_query = request.args.get('q', '')
    query = ItemEstoque.query
    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(or_(
            ItemEstoque.codigo.ilike(search_term),
            ItemEstoque.descricao.ilike(search_term)
        ))
    
    items = query.limit(10).all()
    results = [{'id': item.id, 'text': f"{item.codigo} - {item.descricao}"} for item in items]
    return jsonify(results)

@app.route('/api/item/<int:item_id>/historico-chart')
@login_required
def api_item_historico_chart(item_id):
    """Retorna o histórico de movimentações de um item para o gráfico."""
    hoje = date.today()
    labels_mov = [(hoje - timedelta(days=i)).strftime('%d/%m') for i in range(14, -1, -1)]
    entradas = []
    saidas = []
    for i in range(14, -1, -1):
        dia = hoje - timedelta(days=i)
        entradas.append(db.session.query(func.sum(Movimentacao.quantidade)).filter(Movimentacao.item_id==item_id, Movimentacao.tipo=='ENTRADA', func.date(Movimentacao.data_movimentacao)==dia).scalar() or 0)
        saidas.append(db.session.query(func.sum(Movimentacao.quantidade)).filter(Movimentacao.item_id==item_id, Movimentacao.tipo=='SAIDA', func.date(Movimentacao.data_movimentacao)==dia).scalar() or 0)

    historico_data = {
        'labels': labels_mov,
        'entradas': entradas,
        'saidas': saidas
    }
    return jsonify(historico_data)

@app.route('/api/item/<int:item_id>/previsao')
@login_required
def api_item_previsao(item_id):
    """
    Gera e retorna uma previsão de consumo (saídas) para um item específico.
    Utiliza um modelo de Suavização Exponencial.
    """
    item = ItemEstoque.query.get_or_404(item_id)
    
    # 1. Coletar dados históricos de SAÍDA para o item
    # Busca movimentações dos últimos 90 dias (ajustável)
    data_limite_historico = datetime.now() - timedelta(days=90)
    historico_saidas = Movimentacao.query.filter(
        Movimentacao.item_id == item_id,
        Movimentacao.tipo == 'SAIDA',
        Movimentacao.data_movimentacao >= data_limite_historico
    ).order_by(Movimentacao.data_movimentacao.asc()).all()

    if not historico_saidas:
        return jsonify({'error': 'Dados históricos de saída insuficientes para previsão.', 'previsao': []}), 404

    # 2. Preparar os dados para o modelo (série temporal diária)
    data_saidas = []
    for mov in historico_saidas:
        data_saidas.append({
            'data': mov.data_movimentacao.date(),
            'quantidade': mov.quantidade
        })
    
    df_saidas = pd.DataFrame(data_saidas)
    df_saidas['data'] = pd.to_datetime(df_saidas['data'])
    
    # Agrupar por dia e somar as quantidades
    df_diario = df_saidas.groupby('data')['quantidade'].sum().reset_index()
    df_diario = df_diario.set_index('data')
    
    # Preencher datas ausentes com 0 para ter uma série temporal contínua
    idx = pd.date_range(df_diario.index.min(), df_diario.index.max())
    df_diario = df_diario.reindex(idx, fill_value=0)
    
    # 3. Treinar o modelo de Suavização Exponencial
    # Usando ExponentialSmoothing com tendência aditiva e sem sazonalidade para começar
    try:
        model = ExponentialSmoothing(df_diario['quantidade'], trend='add', seasonal=None, initialization_method="estimated").fit()
        
        # 4. Gerar a previsão para os próximos 15 dias (ajustável)
        dias_para_prever = 15
        forecast_result = model.forecast(dias_para_prever)
        
        previsao_data = []
        for i, value in enumerate(forecast_result):
            forecast_date = df_diario.index.max() + timedelta(days=i+1)
            previsao_data.append({
                'data': forecast_date.strftime('%Y-%m-%d'),
                'quantidade_prevista': max(0, round(value, 2)) # Garante que a previsão não seja negativa
            })
        
        return jsonify({'item_id': item_id, 'item_descricao': item.descricao, 'previsao': previsao_data})
    except Exception as e:
        return jsonify({'error': f'Erro ao gerar previsão: {e}', 'previsao': []}), 500

def _gerar_previsao_para_item(item_id, dias_para_prever):
    """
    Função auxiliar interna para gerar previsão. Não é uma rota.
    Retorna um dicionário com a previsão ou um erro.
    """
    item = ItemEstoque.query.get(item_id)
    if not item:
        return {'error': 'Item não encontrado.'}

    data_limite_historico = datetime.now() - timedelta(days=90)
    historico_saidas = Movimentacao.query.filter(
        Movimentacao.item_id == item_id,
        Movimentacao.tipo == 'SAIDA',
        Movimentacao.data_movimentacao >= data_limite_historico
    ).order_by(Movimentacao.data_movimentacao.asc()).all()

    if not historico_saidas:
        return {'error': 'Dados históricos insuficientes.', 'previsao': []}

    data_saidas = [{'data': mov.data_movimentacao.date(), 'quantidade': mov.quantidade} for mov in historico_saidas]
    df_saidas = pd.DataFrame(data_saidas)
    df_saidas['data'] = pd.to_datetime(df_saidas['data'])
    
    df_diario = df_saidas.groupby('data')['quantidade'].sum().reset_index().set_index('data')
    
    idx = pd.date_range(df_diario.index.min(), df_diario.index.max())
    df_diario = df_diario.reindex(idx, fill_value=0)
    
    try:
        model = ExponentialSmoothing(df_diario['quantidade'], trend='add', seasonal=None, initialization_method="estimated").fit()
        forecast_result = model.forecast(dias_para_prever)
        
        previsao_data = []
        for i, value in enumerate(forecast_result):
            forecast_date = df_diario.index.max() + timedelta(days=i+1)
            previsao_data.append({
                'data': forecast_date.strftime('%Y-%m-%d'),
                'quantidade_prevista': max(0, round(value, 2))
            })
        
        return {'previsao': previsao_data}
    except Exception as e:
        return {'error': f'Erro no modelo de previsão: {e}', 'previsao': []}

@app.route('/api/sugestoes-compra')
@login_required
def api_sugestoes_compra():
    """
    Analisa o estoque, o consumo previsto e o tempo de reposição para gerar
    sugestões de compra inteligentes. Com limite de itens e tratamento de erro.
    """
    try:
        sugestoes = []
        # Analisa apenas itens que têm estoque e com limite para evitar overhead
        itens = ItemEstoque.query.filter(ItemEstoque.qtd_estoque > 0).limit(100).all()

        for item in itens:
            try:
                # 1. Obter a previsão de consumo para o período do lead time
                dias_para_prever = item.tempo_reposicao or 7
                
                # Usa a função auxiliar para obter a previsão
                resultado_previsao = _gerar_previsao_para_item(item.id, dias_para_prever)
                if 'error' in resultado_previsao:
                    continue # Pula para o próximo item se não houver previsão

                previsao_data = resultado_previsao
                consumo_previsto_total = sum(p['quantidade_prevista'] for p in previsao_data.get('previsao', []))

                # 2. Calcular o estoque projetado ao final do lead time
                estoque_projetado = item.qtd_estoque - consumo_previsto_total

                # 3. Verificar se o estoque projetado está abaixo do mínimo
                if estoque_projetado < item.estoque_minimo:
                    # 4. Gerar a sugestão de compra
                    
                    # Usa a quantidade ideal configurada, ou calcula automaticamente
                    if item.estoque_ideal_compra and item.estoque_ideal_compra > 0:
                        quantidade_sugerida = item.estoque_ideal_compra
                    else:
                        # Cálculo automático: para voltar ao dobro do estoque mínimo
                        quantidade_sugerida = (item.estoque_minimo * 2) - estoque_projetado
                    
                    # Calcula até quando o pedido deve ser feito
                    dias_ate_critico = (item.qtd_estoque - item.estoque_minimo) / (consumo_previsto_total / dias_para_prever) if consumo_previsto_total > 0 else float('inf')
                    data_limite_pedido = date.today() + timedelta(days=max(0, dias_ate_critico))

                    sugestoes.append({
                        'item_id': item.id,
                        'item_descricao': item.descricao,
                        'item_codigo': item.codigo,
                        'quantidade_sugerida': round(quantidade_sugerida, 2),
                        'data_limite_pedido': data_limite_pedido.strftime('%d/%m/%Y')
                    })
            except Exception as e:
                # Log do erro e continua para o próximo item
                print(f"Erro ao processar item {item.id}: {str(e)}")
                continue

        # Ordena por data de prazo
        sugestoes_ordenadas = sorted(sugestoes, key=lambda x: datetime.strptime(x['data_limite_pedido'], '%d/%m/%Y'))
        return jsonify(sugestoes_ordenadas)
    
    except Exception as e:
        # Retorna um array vazio com erro, evitando que o frontend fique pendurado
        print(f"Erro geral em api_sugestoes_compra: {str(e)}")
        return jsonify([])

@app.route('/api/stock-turnover-data')
@login_required
def api_stock_turnover_data():
    """
    Calcula e retorna o giro de estoque para identificar itens parados.
    Giro = Total de Saídas (últimos 90 dias) / Estoque Atual
    """
    hoje = date.today()
    periodo_dias = 90 # Período para calcular as saídas
    data_limite_saidas = datetime.now() - timedelta(days=periodo_dias)

    itens_com_giro = []
    # Filtra apenas itens que têm estoque para calcular o giro
    itens = ItemEstoque.query.filter(ItemEstoque.qtd_estoque > 0).all()

    for item in itens:
        total_saidas_periodo = db.session.query(func.sum(Movimentacao.quantidade)).filter(
            Movimentacao.item_id == item.id,
            Movimentacao.tipo == 'SAIDA',
            Movimentacao.data_movimentacao >= data_limite_saidas
        ).scalar() or 0

        giro_estoque = 0
        if item.qtd_estoque > 0: # Evita divisão por zero
            giro_estoque = total_saidas_periodo / item.qtd_estoque
        
        itens_com_giro.append({
            'id': item.id,
            'descricao': item.descricao,
            'codigo': item.codigo, # Mantém o código para o tooltip
            'giro_estoque': round(giro_estoque, 2), # Mantém o código para o tooltip
            'qtd_estoque': item.qtd_estoque,
            'total_saidas': total_saidas_periodo
        })
    
    # Ordena pelo giro de estoque (menor giro = mais parado) e retorna os top 10
    itens_com_giro_ordenado = sorted(itens_com_giro, key=lambda x: x['giro_estoque'])
    return jsonify(itens_com_giro_ordenado[:10])

@app.route('/')
@login_required
def index():
    """Redireciona a rota raiz para o novo dashboard."""
    return redirect(url_for('dashboard'))

@app.route('/cadastro', methods=['GET', 'POST'])
@admin_only
@login_required
def cadastro():
    """Página para cadastrar um novo item no estoque."""
    if request.method == 'POST':
        # Coleta os dados do formulário para ItemEstoque
        codigo = request.form['codigo']
        descricao = request.form['descricao']
        
        # Coleta os dados do formulário para EstoqueDetalhe (primeiro lote)
        qtd_entrada = request.form.get('qtd_estoque') # Renomeado para clareza
        lote = request.form.get('lote')
        item_nf = request.form.get('item_nf')
        nf = request.form.get('nf')
        validade_str = request.form.get('validade')
        # Validação para campos obrigatórios do ItemEstoque
        if not codigo or not descricao:
            flash('Código e Descrição são campos obrigatórios para o item!', 'danger')
            return redirect(url_for('cadastro'))
        
        # Validação para campos obrigatórios do EstoqueDetalhe inicial
        if not qtd_entrada or not lote or not item_nf:
            flash('Quantidade, Lote e Item NF são campos obrigatórios para a entrada inicial!', 'danger')
            return redirect(url_for('cadastro'))

        try:
            qtd_entrada = float(qtd_entrada)
            if qtd_entrada <= 0:
                flash('A quantidade de entrada deve ser um número inteiro positivo.', 'danger')
                return redirect(url_for('cadastro'))
        except (ValueError, TypeError):
            flash('A quantidade de entrada deve ser um número válido.', 'danger')
            return redirect(url_for('cadastro'))

        # Verifica se o item já existe pelo código
        if ItemEstoque.query.filter_by(codigo=codigo).first():
            flash(f'Item com o código "{codigo}" já existe no estoque. Use a tela de movimentação para adicionar mais lotes.', 'warning')
            return redirect(url_for('cadastro'))

        # Converte validade
        validade = datetime.strptime(validade_str, '%Y-%m-%d').date() if validade_str else None

        # Cria um novo objeto ItemEstoque
        novo_item = ItemEstoque(
            codigo=codigo,
            endereco=request.form.get('endereco'),
            codigo_opcional=request.form.get('codigo_opcional'),
            tipo=request.form.get('tipo'),
            descricao=descricao,
            un=request.form.get('un'),
            dimensao=request.form.get('dimensao'),
            estoque_minimo=float(request.form.get('estoque_minimo', 5)), # Adiciona o estoque mínimo
            estoque_ideal_compra=float(request.form.get('estoque_ideal_compra', 0)) if request.form.get('estoque_ideal_compra') else None, # Quantidade ideal para comprar
            cliente=request.form.get('cliente'),
            qtd_estoque=0 # Começa com 0, será atualizado pelo EstoqueDetalhe
        )

        # Adiciona e salva no banco de dados
        try:
            db.session.add(novo_item)
            db.session.flush() # Para obter o ID do novo_item antes de commitar
            
            # Cria o primeiro detalhe de estoque para este item
            novo_detalhe = EstoqueDetalhe(
                item_estoque_id=novo_item.id,
                lote=lote,
                item_nf=item_nf,
                nf=nf,
                validade=validade,
                estacao='Almoxarifado',
                quantidade=qtd_entrada
            )
            db.session.add(novo_detalhe)

            # Atualiza a quantidade total do ItemEstoque
            novo_item.qtd_estoque += qtd_entrada

            # Cria o registro de movimentação para constar no histórico e dashboard
            movimentacao_inicial = Movimentacao(
                item_id=novo_item.id,
                tipo='ENTRADA',
                quantidade=qtd_entrada,
                lote=lote,
                item_nf=item_nf,
                nf=nf, # <-- ADICIONADO: Garante que a NF seja salva na movimentação inicial
                usuario=current_user.username,
                etapa='CADASTRO',
                observacao='Entrada inicial via cadastro de novo item.'
            )
            db.session.add(movimentacao_inicial)

            db.session.commit()
            flash('Item e seu primeiro lote cadastrados com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar o item e lote: {e}', 'danger')
        
        return redirect(url_for('estoque'))

    return render_template('cadastro.html')

@app.route('/controle_validade')
@login_required
def controle_validade():
    """
    Tela de Controle de Validade – Itens para Etiqueta Vermelha (40 dias).
    Exibe itens que estão a 40 dias ou menos do vencimento E com status 'PENDENTE'.
    """
    hoje = date.today()
    data_limite = hoje + timedelta(days=40)
    
    # Busca lotes com validade próxima ou vencida, com quantidade > 0 e PENDENTES
    lotes_criticos = EstoqueDetalhe.query.join(ItemEstoque).filter(
        EstoqueDetalhe.validade <= data_limite,
        EstoqueDetalhe.validade.isnot(None),
        EstoqueDetalhe.quantidade > 0,
        or_(EstoqueDetalhe.status_etiqueta == 'PENDENTE', EstoqueDetalhe.status_etiqueta == None)
    ).order_by(EstoqueDetalhe.validade.asc()).all()
    
    return render_template('controle_validade.html', 
                           lotes=lotes_criticos, 
                           today_date=hoje, 
                           title="Controle de Validade")

@app.route('/controle_validade/marcar_concluido/<int:lote_id>', methods=['POST'])
@login_required
def marcar_validade_concluido(lote_id):
    """
    Marca um item como 'CONCLUÍDO' no processo de etiqueta vermelha.
    """
    lote = EstoqueDetalhe.query.get_or_404(lote_id)
    
    try:
        lote.status_etiqueta = 'CONCLUÍDO'
        lote.data_etiqueta = datetime.now()
        lote.usuario_etiqueta = current_user.username
        db.session.commit()
        return jsonify({'success': True, 'message': 'Item marcado como concluído!'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/controle_validade/historico')
@login_required
def historico_validade():
    """
    Exibe o histórico de itens já etiquetados (CONCLUÍDO).
    """
    hoje = date.today()
    
    lotes_concluidos = EstoqueDetalhe.query.join(ItemEstoque).filter(
        EstoqueDetalhe.status_etiqueta == 'CONCLUÍDO'
    ).order_by(EstoqueDetalhe.data_etiqueta.desc()).limit(500).all() # Limitando para não pesar
    
    return render_template('controle_validade_historico.html', 
                           lotes=lotes_concluidos, 
                           today_date=hoje, 
                           title="Histórico de Etiquetas")

@app.route('/controle_validade/reabrir/<int:lote_id>', methods=['POST'])
@admin_only
@login_required
def reabrir_validade(lote_id):
    """
    Reabre um item concluído, voltando para PENDENTE (Admin only).
    """
    lote = EstoqueDetalhe.query.get_or_404(lote_id)
    
    try:
        lote.status_etiqueta = 'PENDENTE'
        # Mantemos o histórico de data/usuário ou limpamos? Melhor limpar para indicar novo ciclo.
        lote.data_etiqueta = None
        lote.usuario_etiqueta = None
        db.session.commit()
        flash('Item reaberto com sucesso. Ele voltou para a lista de pendências.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reabrir item: {e}', 'danger')
        
    return redirect(url_for('historico_validade'))

@app.route('/controle_validade/exportar')
@admin_only
@login_required
def exportar_validade():
    """
    Exporta a lista de controle de validade para Excel.
    Se passar ?tipo=historico, exporta o histórico.
    """
    tipo = request.args.get('tipo', 'pendente')
    hoje = date.today()
    output = io.BytesIO()
    
    # Query base
    query = EstoqueDetalhe.query.join(ItemEstoque).filter(
        EstoqueDetalhe.validade.isnot(None),
        EstoqueDetalhe.quantidade > 0
    )
    
    if tipo == 'historico':
        query = query.filter(EstoqueDetalhe.status_etiqueta == 'CONCLUÍDO').order_by(EstoqueDetalhe.data_etiqueta.desc())
        filename = f"Historico_Etiquetas_{hoje.strftime('%Y-%m-%d')}.xlsx"
    else:
        # Padrão: Pendentes (<= 40 dias)
        data_limite = hoje + timedelta(days=40)
        query = query.filter(
            EstoqueDetalhe.validade <= data_limite,
            or_(EstoqueDetalhe.status_etiqueta == 'PENDENTE', EstoqueDetalhe.status_etiqueta == None)
        ).order_by(EstoqueDetalhe.validade.asc())
        filename = f"Pendencias_Etiquetas_{hoje.strftime('%Y-%m-%d')}.xlsx"
        
    lotes = query.all()
    
    data = []
    for lote in lotes:
        dias_para_vencer = (lote.validade - hoje).days
        status_vencimento = "VENCIDO" if dias_para_vencer < 0 else "Crítico"
        
        row = {
            'Código': lote.item_estoque.codigo,
            'Descrição': lote.item_estoque.descricao,
            'Lote': lote.lote,
            'Local': f"{lote.item_estoque.endereco or '-'} / {lote.estacao or '-'}",
            'Validade': lote.validade.strftime('%d/%m/%Y'),
            'Dias Vencimento': dias_para_vencer,
            'Quantidade': lote.quantidade,
            'Status Etiqueta': lote.status_etiqueta or 'PENDENTE'
        }
        
        if tipo == 'historico':
            row['Data Etiqueta'] = lote.data_etiqueta.strftime('%d/%m/%Y %H:%M') if lote.data_etiqueta else '-'
            row['Resp. Etiqueta'] = lote.usuario_etiqueta or '-'
            
        data.append(row)
        
    df = pd.DataFrame(data)
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Controle Validade')
        workbook = writer.book
        worksheet = writer.sheets['Controle Validade']
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
        for i, col in enumerate(df.columns):
            worksheet.set_column(i, i, 20)
            worksheet.write(0, i, col, header_format)
            
    output.seek(0)
    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response

@app.route('/controle_validade/imprimir')
@admin_only
@login_required
def imprimir_validade():
    """
    Renderiza a versão de impressão da lista de controle de validade (PENDENTES).
    """
    hoje = date.today()
    data_limite = hoje + timedelta(days=40)
    
    lotes_criticos = EstoqueDetalhe.query.join(ItemEstoque).filter(
        EstoqueDetalhe.validade <= data_limite,
        EstoqueDetalhe.validade.isnot(None),
        EstoqueDetalhe.quantidade > 0,
        or_(EstoqueDetalhe.status_etiqueta == 'PENDENTE', EstoqueDetalhe.status_etiqueta == None)
    ).order_by(EstoqueDetalhe.validade.asc()).all()
    
    return render_template('print_validade.html', 
                           lotes=lotes_criticos, 
                           today_date=hoje)

@app.route('/estoque')
@login_required
def estoque():
    """
    Página para listar todos os lotes detalhados do estoque.
    A exibição é baseada em EstoqueDetalhe, mostrando cada lote/nf como uma linha.
    """
    search_query = request.args.get('q', '')
    
    # A consulta base agora é em EstoqueDetalhe, com join para ItemEstoque para busca e exibição.
    query = EstoqueDetalhe.query.join(ItemEstoque, EstoqueDetalhe.item_estoque_id == ItemEstoque.id)
    
    if search_query and search_query.strip():
        search_term = f"%{search_query}%"
        # O filtro agora busca em campos do ItemEstoque e do EstoqueDetalhe
        query = query.filter(
            or_(
                ItemEstoque.codigo.ilike(search_term),
                ItemEstoque.descricao.ilike(search_term),
                EstoqueDetalhe.lote.ilike(search_term),
                EstoqueDetalhe.nf.ilike(search_term),
                EstoqueDetalhe.item_nf.ilike(search_term)
            )
        )
    
    # Filtra apenas lotes que têm saldo positivo
    query = query.filter(EstoqueDetalhe.quantidade > 0)
    
    # Ordena por FIFO (primeiro que vence, primeiro que sai). Lotes sem validade ficam por último.
    query = query.order_by(EstoqueDetalhe.validade.asc().nullslast(), ItemEstoque.codigo.asc())
    
    # Implementação da Paginação sobre a nova query
    page = request.args.get('page', 1, type=int)
    per_page = 25
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    
    # A variável agora contém objetos EstoqueDetalhe paginados
    lotes_detalhados = pagination.items
    
    # O template receberá 'lotes_detalhados' em vez de 'itens'
    return render_template('estoque.html',
                           lotes_detalhados=lotes_detalhados,
                           search_query=search_query,
                           pagination=pagination,
                           today_date=date.today(),
                           timedelta=timedelta)

@app.route('/item/editar/<int:item_id>', methods=['GET', 'POST'])
@admin_only
@login_required
def editar_item(item_id):
    """Página para editar um item existente."""
    item = ItemEstoque.query.get_or_404(item_id)

    if request.method == 'POST':
        # Coleta os dados do formulário
        codigo = request.form['codigo']
        descricao = request.form['descricao']

        # Validação para campos obrigatórios
        if not codigo or not descricao:
            flash('Código e Descrição são campos obrigatórios!', 'danger')
            return render_template('editar_item.html', item=item)

        # Verifica se o novo código já existe em outro item
        item_existente = ItemEstoque.query.filter(ItemEstoque.codigo == codigo, ItemEstoque.id != item_id).first()
        if item_existente:
            flash(f'O código "{codigo}" já está em uso por outro item.', 'warning')
            return render_template('editar_item.html', item=item)

        # Atualiza os campos do objeto item
        item.codigo = codigo
        item.descricao = descricao
        item.endereco = request.form.get('endereco')
        item.codigo_opcional = request.form.get('codigo_opcional')
        item.tipo = request.form.get('tipo')
        item.un = request.form.get('un')
        item.dimensao = request.form.get('dimensao')
        item.estoque_minimo = float(request.form.get('estoque_minimo', 5)) # Atualiza o estoque mínimo
        item.estoque_ideal_compra = float(request.form.get('estoque_ideal_compra', 0)) if request.form.get('estoque_ideal_compra') else None # Atualiza a quantidade ideal para comprar
        item.tempo_reposicao = int(request.form.get('tempo_reposicao', 7)) # Atualiza o tempo de reposição
        item.cliente = request.form.get('cliente')
        # Os campos de lote, NF, validade, estacao, status_validade e qtd_estoque
        # NÃO são mais parte direta do ItemEstoque e não devem ser atualizados aqui.
        # Eles pertencem ao EstoqueDetalhe.
        try:
            db.session.commit()
            flash('Item atualizado com sucesso!', 'success')
            return redirect(url_for('estoque'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar o item: {e}', 'danger')

    return render_template('editar_item.html', item=item)

@app.route('/item/excluir/<int:item_id>')
@admin_required
@login_required
def excluir_item(item_id):
    """Rota para excluir um item do estoque."""
    item = ItemEstoque.query.get_or_404(item_id)
    # Adicionar verificação se o item tem movimentações? (Opcional)
    db.session.delete(item)
    db.session.commit()
    flash(f'Item "{item.descricao}" e todo o seu histórico foram excluídos com sucesso.', 'success')
    return redirect(url_for('estoque'))

@app.route('/movimentacao', methods=['GET', 'POST'])
@admin_only
@login_required
def movimentacao():
    """Página para registrar uma nova movimentação de estoque."""
    if request.method == 'POST':
        item_id = request.form.get('item_id')
        tipo = request.form.get('tipo')
        etapa = request.form.get('etapa') # Captura a etapa
        observacao = request.form.get('observacao') # Captura a observação
        
        if not item_id or not tipo:
            flash('Item ou tipo de movimentação inválido.', 'danger')
            return redirect(url_for('movimentacao'))

        item = ItemEstoque.query.get_or_404(item_id)

        if not etapa:
            flash('O campo Etapa de Destino é obrigatório.', 'danger')
            return redirect(url_for('movimentacao'))

        try:
            quantidade = float(request.form.get('quantidade'))
            if quantidade <= 0:
                flash('A quantidade deve ser um número positivo.', 'danger')
                return redirect(url_for('movimentacao'))
        except (ValueError, TypeError):
            flash('A quantidade deve ser um número válido.', 'danger')
            return redirect(url_for('movimentacao'))

        # --- LÓGICA DE ENTRADA ---
        if tipo == 'ENTRADA':
            lote = request.form.get('lote')
            item_nf = request.form.get('item_nf')
            nf = request.form.get('nf') # Captura a NF para uso em todo o bloco
            if not lote or not item_nf:
                flash('Lote e Item NF são obrigatórios para entrada.', 'danger')
                return redirect(url_for('movimentacao'))

            # Verifica se já existe um detalhe com a mesma combinação de lote, item_nf e nf
            detalhe_existente = item.detalhes_estoque.filter_by(
                lote=lote, 
                item_nf=item_nf, 
                nf=nf
            ).first()
            
            if detalhe_existente:
                # Se existe, apenas soma a quantidade
                detalhe_existente.quantidade += quantidade
            else:
                # Se não existe, cria um novo registro de EstoqueDetalhe
                validade_str = request.form.get('validade')

                # Lógica para determinar a estação automaticamente
                # Busca o último lote inserido para este item
                ultimo_detalhe = item.detalhes_estoque.order_by(EstoqueDetalhe.data_entrada.desc()).first()
                # Usa a estação do último lote como padrão, ou 'Almoxarifado' se não houver lote anterior ou se a estação do último lote for nula
                estacao_automatica = ultimo_detalhe.estacao if ultimo_detalhe and ultimo_detalhe.estacao else 'Almoxarifado'

                novo_detalhe = EstoqueDetalhe(
                    item_estoque_id=item.id,
                    lote=lote,
                    item_nf=item_nf,
                    nf=nf,
                    validade=datetime.strptime(validade_str, '%Y-%m-%d').date() if validade_str else None,
                    estacao=estacao_automatica,
                    quantidade=quantidade
                )
                db.session.add(novo_detalhe)

            # Atualiza o total e registra a movimentação
            item.qtd_estoque += quantidade
            nova_movimentacao = Movimentacao(
                item_id=item.id, 
                tipo='ENTRADA', 
                quantidade=quantidade, 
                lote=lote, 
                item_nf=item_nf,
                nf=nf,
                usuario=current_user.username,
                etapa=etapa,
                observacao=observacao)
            db.session.add(nova_movimentacao)
            db.session.commit()
            socketio.emit('dashboard_update', {'message': 'Nova entrada registrada!'})
            flash('Entrada registrada com sucesso!', 'success')
        # --- LÓGICA DE SAÍDA ---
        elif tipo == 'SAIDA':
            detalhe_id = request.form.get('detalhe_id')
            if not detalhe_id:
                flash('É necessário selecionar um Lote/Item NF para a saída.', 'danger')
                return redirect(url_for('movimentacao'))

            detalhe_estoque = EstoqueDetalhe.query.get(detalhe_id)

            if not detalhe_estoque or detalhe_estoque.item_estoque_id != item.id:
                flash('Lote/Item NF inválido para este item.', 'danger')
                return redirect(url_for('movimentacao'))

            if detalhe_estoque.quantidade < quantidade:
                flash(f'Quantidade insuficiente no lote selecionado. Disponível: {detalhe_estoque.quantidade}', 'danger')
                return redirect(url_for('movimentacao'))

            # Subtrai a quantidade do lote específico e do total
            detalhe_estoque.quantidade -= quantidade
            item.qtd_estoque -= quantidade

            # Registra a movimentação
            nova_movimentacao = Movimentacao(
                item_id=item.id, 
                tipo='SAIDA', 
                quantidade=quantidade, 
                lote=detalhe_estoque.lote, 
                item_nf=detalhe_estoque.item_nf,
                nf=detalhe_estoque.nf,
                usuario=current_user.username,
                etapa=etapa,
                observacao=observacao)
            db.session.add(nova_movimentacao)
            db.session.commit()
            socketio.emit('dashboard_update', {'message': 'Nova saída registrada!'})
            flash('Saída registrada com sucesso!', 'success')
        return redirect(url_for('movimentacao'))

    # Para o método GET (carregamento da página)
    return render_template('movimentacao.html')

@app.route('/historico/<int:item_id>')
@admin_only
@login_required
def historico(item_id):
    """Página para exibir o histórico de movimentações de um item."""
    item = ItemEstoque.query.get_or_404(item_id)
    # Ordena as movimentações da mais recente para a mais antiga
    movimentacoes = item.movimentacoes.order_by(Movimentacao.data_movimentacao.desc()).all()
    return render_template('historico.html', item=item, movimentacoes=movimentacoes)

@app.route('/item/<int:item_id>/lotes-detalhes')
@login_required
def detalhes_lotes(item_id):
    """Página para exibir todos os detalhes de lote de um item."""
    filtro_ativo = request.args.get('filtro')
    item = ItemEstoque.query.get_or_404(item_id)
    query = item.detalhes_estoque

    if filtro_ativo == 'critico':
        # Define a data limite: hoje + 30 dias
        data_limite = date.today() + timedelta(days=30)
        # Filtra lotes que têm validade e que vencem até a data limite
        query = query.filter(
            EstoqueDetalhe.quantidade > 0,  # Apenas lotes com saldo
            EstoqueDetalhe.validade.isnot(None),
            EstoqueDetalhe.validade <= data_limite
        )

    # Ordena os resultados pela data de entrada para ver os mais recentes primeiro
    detalhes_query = query.order_by(
        EstoqueDetalhe.data_entrada.desc()
    ).all()

    return render_template('detalhes_lotes.html',
                         item=item,
                         detalhes=detalhes_query,
                         filtro_ativo=filtro_ativo)
 
@app.route('/lote/editar/<int:detalhe_id>', methods=['GET', 'POST'])
@admin_required
@login_required
def editar_lote(detalhe_id):
    """Página para editar os dados de um lote específico (saldo, NF, validade, etc.)."""
    detalhe = EstoqueDetalhe.query.get_or_404(detalhe_id)
    item = detalhe.item_estoque

    if request.method == 'POST':
        try:
            # --- Coleta e Validação dos Dados ---
            nova_quantidade_str = request.form.get('quantidade')
            observacao = request.form.get('observacao')
            novo_endereco = request.form.get('endereco') # <-- Adicionado para capturar o endereço

            if not nova_quantidade_str or not observacao:
                flash('A Nova Quantidade e o Motivo da Edição são obrigatórios.', 'danger')
                return render_template('editar_lote.html', detalhe=detalhe)

            nova_quantidade = float(nova_quantidade_str)
            if nova_quantidade < 0:
                flash('A quantidade não pode ser negativa.', 'danger')
                return render_template('editar_lote.html', detalhe=detalhe)

            # --- Cálculo do Ajuste de Quantidade ---
            quantidade_antiga = detalhe.quantidade
            diferenca_qtd = nova_quantidade - quantidade_antiga

            # --- Atualização dos Dados do Lote (EstoqueDetalhe) ---
            detalhe.lote = request.form.get('lote') # <-- Adicionado para capturar o Lote
            detalhe.nf = request.form.get('nf')
            detalhe.item_nf = request.form.get('item_nf')
            validade_str = request.form.get('validade')
            detalhe.validade = datetime.strptime(validade_str, '%Y-%m-%d').date() if validade_str else None
            detalhe.estacao = request.form.get('estacao') # <-- Captura o valor do campo de texto da estação
            
            # Atualiza o endereço no item principal (ItemEstoque)
            item.endereco = novo_endereco
            detalhe.quantidade = nova_quantidade

            # --- Atualização do Estoque Total do Item (ItemEstoque) ---
            item.qtd_estoque += diferenca_qtd

            # --- Registro da Movimentação de Ajuste ---
            # Define o tipo de ajuste (entrada ou saída) para o histórico
            tipo_ajuste = 'AJUSTE-ENTRADA' if diferenca_qtd > 0 else 'AJUSTE-SAIDA'
            
            # Garante que a quantidade na movimentação seja sempre positiva
            qtd_movimentacao = abs(diferenca_qtd)

            # Só registra a movimentação se houver mudança na quantidade
            if qtd_movimentacao > 0:
                ajuste_movimentacao = Movimentacao(
                    item_id=item.id,
                    tipo=tipo_ajuste,
                    quantidade=qtd_movimentacao,
                    lote=detalhe.lote,
                    item_nf=detalhe.item_nf,
                    nf=detalhe.nf,
                    usuario=current_user.username,
                    etapa='AJUSTE',
                    observacao=f"Ajuste manual. Motivo: {observacao}. Qtd anterior: {quantidade_antiga}, Qtd nova: {nova_quantidade}."
                )
                db.session.add(ajuste_movimentacao)

            db.session.commit()
            flash('Lote editado com sucesso! O histórico de movimentação foi atualizado.', 'success')
            return redirect(url_for('detalhes_lotes', item_id=item.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Ocorreu um erro ao tentar editar o lote: {e}', 'danger')

    return render_template('editar_lote.html', detalhe=detalhe)

def _calcular_quantidade_recebida(detalhe_lote):
    """
    Calcula a quantidade total recebida para um lote específico (detalhe).
    Soma todas as movimentações de ENTRADA e AJUSTE-ENTRADA.
    """
    total_recebido = db.session.query(func.sum(Movimentacao.quantidade)).filter(
        Movimentacao.item_id == detalhe_lote.item_estoque_id,
        Movimentacao.lote == detalhe_lote.lote,
        Movimentacao.nf == detalhe_lote.nf, # Adicionado para garantir a unicidade
        or_(
            Movimentacao.tipo == 'ENTRADA',
            Movimentacao.tipo == 'AJUSTE-ENTRADA'
        )
    ).scalar()

    return total_recebido or 0

@app.route('/lote/excluir/<int:detalhe_id>')
@admin_required
@login_required
def excluir_lote(detalhe_id):
    """
    Exclui um lote (EstoqueDetalhe) específico do estoque.
    Esta ação também remove as movimentações relacionadas e ajusta o saldo do item.
    Esta ação é restrita a administradores.
    """
    detalhe = EstoqueDetalhe.query.get_or_404(detalhe_id)
    item = detalhe.item_estoque
    
    try:
        # Armazena a quantidade para ajuste do estoque
        quantidade_lote = detalhe.quantidade
        
        # Remove todas as movimentações relacionadas a este lote
        movimentacoes_lote = Movimentacao.query.filter_by(item_id=item.id, lote=detalhe.lote, nf=detalhe.nf).all()
        
        for mov in movimentacoes_lote:
            db.session.delete(mov)
        
        # Ajusta o estoque total do item
        item.qtd_estoque -= quantidade_lote
        
        # Garante que o estoque não fique negativo
        if item.qtd_estoque < 0:
            item.qtd_estoque = 0
        
        # Remove o detalhe do estoque
        db.session.delete(detalhe)
        db.session.commit()
        
        flash(f'Lote "{detalhe.lote} / {detalhe.item_nf}" foi excluído com sucesso! Todas as movimentações relacionadas foram removidas.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir o lote: {e}', 'danger')
    
    return redirect(url_for('detalhes_lotes', item_id=item.id))

@app.route('/movimentacao/excluir/<int:mov_id>')
@admin_required
@login_required
def excluir_movimentacao(mov_id):
    """
    Exclui (reverte) uma movimentação de estoque.
    Esta ação é restrita a administradores.
    """
    mov = Movimentacao.query.get_or_404(mov_id)
    item = mov.item
    quantidade_mov = mov.quantidade

    try:
        # Encontra o lote específico relacionado à movimentação
        detalhe_lote = EstoqueDetalhe.query.filter_by(item_estoque_id=item.id, lote=mov.lote, nf=mov.nf).first()

        # Lógica de reversão
        if 'ENTRADA' in mov.tipo:
            # Se está revertendo uma ENTRADA, subtrai do estoque
            item.qtd_estoque -= quantidade_mov
            if detalhe_lote:
                detalhe_lote.quantidade -= quantidade_mov
        elif 'SAIDA' in mov.tipo:
            # Se está revertendo uma SAÍDA, adiciona de volta ao estoque
            item.qtd_estoque += quantidade_mov
            if detalhe_lote:
                detalhe_lote.quantidade += quantidade_mov
            else:
                # Caso raro onde o lote foi deletado, não podemos reverter no detalhe
                flash('Aviso: O lote original desta movimentação não foi encontrado. O estoque total do item foi ajustado, mas o saldo do lote não.', 'warning')

        # Após ajustar os saldos, remove o registro da movimentação
        db.session.delete(mov)
        db.session.commit()
        flash(f'Movimentação ID {mov.id} foi revertida com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao reverter a movimentação: {e}', 'danger')

    # Redireciona de volta para a página de onde o usuário veio (ou para o relatório padrão)
    return redirect(request.referrer or url_for('relatorio_movimentacoes'))


@app.route('/movimentacao/apagar/<int:mov_id>')
@admin_required
@login_required
def apagar_movimentacao(mov_id):
    """
    Apaga um registro de movimentação SEM afetar o estoque.
    Esta ação é restrita a administradores.
    """
    mov = Movimentacao.query.get_or_404(mov_id)

    try:
        # Simplesmente remove o registro da movimentação sem reverter estoque
        db.session.delete(mov)
        db.session.commit()
        flash(f'Linha de movimentação ID {mov.id} foi apagada com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao apagar a movimentação: {e}', 'danger')

    # Redireciona de volta para a página de onde o usuário veio (ou para o relatório padrão)
    return redirect(request.referrer or url_for('relatorio_movimentacoes'))

# --- ROTAS DE API INTERNA (para JavaScript) ---

@app.route('/api/item/by-code/<string:codigo>')
@login_required
def api_get_item_by_code(codigo):
    """Retorna os dados de um item pelo seu código em formato JSON."""
    item = ItemEstoque.query.filter_by(codigo=codigo).first()
    if not item:
        # Se o item não for encontrado, retorna um erro 404 claro.
        return jsonify({'error': 'Item não encontrado'}), 404
    
    # Reutiliza a função api_get_item para não duplicar o código de formatação do JSON.
    return api_get_item(item.id)

@app.route('/api/item/<int:item_id>')
@login_required
def api_get_item(item_id):
    """Retorna os dados de um item em formato JSON."""
    item = ItemEstoque.query.get_or_404(item_id)
    return jsonify({
        'id': item.id,
        'codigo': item.codigo,
        'descricao': item.descricao,
        'un': item.un,
        'dimensao': item.dimensao,
        'cliente': item.cliente
    })

@app.route('/api/item/<int:item_id>/lotes')
@login_required
def api_get_lotes(item_id):
    """
    Retorna os lotes disponíveis para um item em formato JSON, ordenados por FEFO ou FIFO.
    - FEFO (First-Expired, First-Out): Para itens perecíveis, ordena por data de validade.
    - FIFO (First-In, First-Out): Para 'Hardware' e 'Painel', ordena por data de entrada.
    """
    item = ItemEstoque.query.get_or_404(item_id)
    
    query = item.detalhes_estoque.filter(EstoqueDetalhe.quantidade > 0)

    # Verifica o tipo do item para decidir a estratégia de ordenação
    if item.tipo and item.tipo.lower() in ['hardware', 'painel']:
        # FIFO: Primeiro que Entra, Primeiro que Sai (ordena pela data de entrada mais antiga)
        detalhes = query.order_by(EstoqueDetalhe.data_entrada.asc()).all()
    else:
        # FEFO: Primeiro que Vence, Primeiro que Sai (ordena pela data de validade mais próxima)
        # Lotes sem validade ou com validade nula são colocados no final da lista.
        detalhes = query.order_by(EstoqueDetalhe.validade.asc().nullslast(), EstoqueDetalhe.data_entrada.asc()).all()

    lotes = [{'id': d.id, 'lote': d.lote, 'item_nf': d.item_nf, 'quantidade': d.quantidade, 'validade': d.validade.strftime('%d/%m/%Y') if d.validade else 'N/A'} for d in detalhes]
    return jsonify(lotes)

@app.route('/importar', methods=['GET', 'POST'])
@admin_required
@login_required
def importar():
    """Página e lógica para importar dados de uma planilha Excel."""
    if request.method == 'POST':
        # Verifica se o arquivo foi enviado
        if 'arquivo_excel' not in request.files:
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)
        
        file = request.files['arquivo_excel']

        # Verifica se o nome do arquivo está vazio
        if file.filename == '':
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file)
                
                # Validação de colunas obrigatórias
                required_columns = ['CÓDIGO', 'DESCRIÇÃO', 'LOTE', 'NF', 'QTD ESTOQUE']
                if not all(col in df.columns for col in required_columns):
                    flash(f'A planilha deve conter as colunas obrigatórias: {", ".join(required_columns)}', 'danger')
                    return redirect(request.url)

                sucesso_count = 0
                erro_count = 0
                
                for index, row in df.iterrows():
                    # Pega os dados da linha, tratando valores nulos (NaN) do pandas
                    codigo = str(row['CÓDIGO']) if pd.notna(row['CÓDIGO']) else None
                    descricao = str(row['DESCRIÇÃO']) if pd.notna(row['DESCRIÇÃO']) else 'Sem Descrição'
                    # Se LOTE ou NF estiverem vazios, usa 'N/A' como padrão
                    lote = str(row['LOTE']) if pd.notna(row['LOTE']) else 'N/A'
                    nf = str(row.get('NF')) if pd.notna(row.get('NF')) else 'N/A'
                    qtd_entrada = float(row['QTD ESTOQUE']) if pd.notna(row['QTD ESTOQUE']) else 0

                    # Pula a linha APENAS se o CÓDIGO estiver faltando.
                    # Permite a importação de itens com quantidade zero, se necessário.
                    if not codigo:
                        erro_count += 1
                        continue

                    # Procura ou cria o ItemEstoque (item principal)
                    item = ItemEstoque.query.filter_by(codigo=codigo).first()
                    if not item:
                        item = ItemEstoque(
                            codigo=codigo,
                            descricao=descricao,
                            endereco=str(row.get('LOCAL', '')) if pd.notna(row.get('LOCAL')) else '',
                            codigo_opcional=str(row.get('CÓDIGO OPCIONAL', '')) if pd.notna(row.get('CÓDIGO OPCIONAL')) else '',
                            tipo=str(row.get('TIPO', '')) if pd.notna(row.get('TIPO')) else '',
                            un=str(row.get('UN.', '')) if pd.notna(row.get('UN.')) else '',
                            dimensao=str(row.get('DIMENSÃO', '')) if pd.notna(row.get('DIMENSÃO')) else '',
                            cliente=str(row.get('CLIENTE', '')) if pd.notna(row.get('CLIENTE')) else '',
                            qtd_estoque=0
                        )
                        db.session.add(item)
                        db.session.flush() # Para obter o ID do item

                    # Procura ou cria o EstoqueDetalhe (lote) usando a chave composta lote + item_nf + nf
                    item_nf = str(row.get('ITEM NF')) if pd.notna(row.get('ITEM NF')) else 'N/A'
                    
                    detalhe_existente = item.detalhes_estoque.filter_by(lote=lote, item_nf=item_nf, nf=nf).first()
                    
                    if detalhe_existente:
                        # Se a combinação exata já existe, soma a quantidade
                        detalhe_existente.quantidade += qtd_entrada
                    else:
                        # Se não existe, cria um novo registro de lote detalhado
                        validade = pd.to_datetime(row.get('VALIDADE')).date() if 'VALIDADE' in row and pd.notna(row.get('VALIDADE')) else None
                        novo_detalhe = EstoqueDetalhe(
                            item_estoque_id=item.id,
                            lote=lote,
                            item_nf=item_nf,
                            nf=nf,
                            validade=validade,
                            estacao=str(row.get('ESTAÇÃO', '')) if pd.notna(row.get('ESTAÇÃO')) else '',
                            quantidade=qtd_entrada
                        )
                        db.session.add(novo_detalhe)
                    
                    # Atualiza o total e registra a movimentação
                    item.qtd_estoque += qtd_entrada
                    nova_movimentacao = Movimentacao(
                        item_id=item.id, 
                        tipo='ENTRADA', 
                        quantidade=qtd_entrada, 
                        lote=lote, 
                        nf=nf,
                        item_nf=str(row.get('ITEM NF', 'N/A')), 
                        usuario=current_user.username)
                    db.session.add(nova_movimentacao)
                    sucesso_count += 1

                db.session.commit()
                flash(f'Importação concluída! {sucesso_count} registros processados com sucesso. {erro_count} linhas ignoradas por falta de dados.', 'success')

            except Exception as e:
                db.session.rollback()
                flash(f'Ocorreu um erro ao processar o arquivo: {e}', 'danger')
                return redirect(request.url)

            return redirect(url_for('estoque'))

        else:
            flash('Formato de arquivo inválido. Por favor, envie um arquivo .xlsx', 'danger')
            return redirect(request.url)

    return render_template('importar.html')

@app.route('/exportar/excel')
@login_required
def exportar_excel():
    """
    Exporta um relatório detalhado do estoque para um arquivo Excel.
    
    MAPEAMENTO CORRETO (CRÍTICO):
    Definido explicitamente conforme solicitação para garantir integridade.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # DEFINIÇÃO EXPLÍCITA DA ORDEM DE COLUNAS
        # Esta lista define EXATAMENTE qual coluna recebe qual dado
        colunas_ordem = [
            'CÓDIGO', 'CÓDIGO OPCIONAL', 'TIPO', 'DESCRIÇÃO',
            'LOCAL', 'UN', 'DIMENSÃO', 'CLIENTE',
            'LOTE', 'ITEM NF', 'NF',
            'VALIDADE', 'ESTAÇÃO', 'QTD ESTOQUE', 'DATA ENTRADA'
        ]
        
        # Busca todos os detalhes de lote com join no ItemEstoque
        # Ordenação sugerida: Código e Data de Entrada
        query = db.session.query(EstoqueDetalhe).join(ItemEstoque).order_by(
            ItemEstoque.codigo, 
            EstoqueDetalhe.data_entrada
        ).all()

        # Cria um workbook novo
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Estoque_Detalhado'
        
        # ========== ESCREVER CABEÇALHO (LINHA 1) ==========
        for num_coluna, nome_coluna in enumerate(colunas_ordem, start=1):
            célula = worksheet.cell(row=1, column=num_coluna)
            célula.value = nome_coluna
            célula.fill = PatternFill(start_color='00D4FF', end_color='00D4FF', fill_type='solid')
            célula.font = Font(bold=True, color='FFFFFF', size=11)
            célula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # ========== ESCREVER DADOS (LINHAS 2+) ==========
        for num_linha, detalhe in enumerate(query, start=2):
            item = detalhe.item_estoque
            
            # Montagem Explícita da Linha (row) conforme solicitado
            dados_linha = {
                'CÓDIGO': str(item.codigo).strip() if item.codigo else '',
                'CÓDIGO OPCIONAL': str(item.codigo_opcional).strip() if item.codigo_opcional else '',
                'TIPO': str(item.tipo).strip() if item.tipo else '',
                'DESCRIÇÃO': str(item.descricao).strip() if item.descricao else '',
                'LOCAL': str(item.endereco).strip() if item.endereco else '',
                'UN': str(item.un).strip() if item.un else 'UN',
                'DIMENSÃO': str(item.dimensao).strip() if item.dimensao else '',
                'CLIENTE': str(item.cliente).strip() if item.cliente else '',
                'LOTE': str(detalhe.lote).strip() if detalhe.lote else '',
                'ITEM NF': str(detalhe.item_nf).strip() if detalhe.item_nf else '',
                'NF': str(detalhe.nf).strip() if detalhe.nf else '',
                'VALIDADE': detalhe.validade.strftime('%d/%m/%Y') if detalhe.validade else '',
                'ESTAÇÃO': str(detalhe.estacao).strip() if detalhe.estacao else '',
                'QTD ESTOQUE': round(float(detalhe.quantidade), 2) if detalhe.quantidade else 0,
                'DATA ENTRADA': detalhe.data_entrada.strftime('%d/%m/%Y %H:%M:%S') if detalhe.data_entrada else ''
            }
            
            # Escrever cada valor na coluna correta
            for num_coluna, nome_coluna in enumerate(colunas_ordem, start=1):
                célula = worksheet.cell(row=num_linha, column=num_coluna)
                valor = dados_linha.get(nome_coluna, '')
                célula.value = valor
                célula.alignment = Alignment(horizontal='left', vertical='center')
        
        # ========== AJUSTE LEVE DE LARGURA (Opcional, para visualização) ==========
        for num_coluna, nome_coluna in enumerate(colunas_ordem, start=1):
            letra_coluna = get_column_letter(num_coluna)
            if nome_coluna in ['DESCRIÇÃO']:
                worksheet.column_dimensions[letra_coluna].width = 40
            elif nome_coluna in ['LOCAL', 'CLIENTE', 'TIPO']:
                worksheet.column_dimensions[letra_coluna].width = 20
            else:
                worksheet.column_dimensions[letra_coluna].width = 15
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        response = make_response(output.read())
        response.headers["Content-Disposition"] = f"attachment; filename=relatorio_estoque_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        return response

    except Exception as e:
        flash(f'Erro ao gerar o relatório: {str(e)}', 'danger')
        return redirect(url_for('estoque'))

@app.route('/estoque/apagar-tudo')
@admin_required
@login_required
def limpar_estoque_completo():
    """Apaga todos os registros de estoque para iniciar um novo inventário."""
    try:
        # A ordem é importante: primeiro os 'filhos', depois os 'pais'
        db.session.query(Movimentacao).delete()
        db.session.query(EstoqueDetalhe).delete()
        db.session.query(ItemEstoque).delete()
        db.session.commit()
        flash('TODO O ESTOQUE FOI APAGADO COM SUCESSO! Agora você pode importar uma nova planilha de inventário.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ocorreu um erro ao tentar limpar o estoque: {e}', 'danger')
    
    return redirect(url_for('estoque'))

# --- ROTAS DE GERENCIAMENTO DE CONSUMÍVEIS ---

@app.route('/consumivel')
@login_required
def consumivel():
    """Página para listar todos os itens consumíveis do estoque."""
    search_query = request.args.get('q', '')
    
    query = ConsumivelEstoque.query

    if search_query:
        search_term = f"%{search_query}%"
        query = query.filter(or_(
            ConsumivelEstoque.codigo_produto.ilike(search_term),
            ConsumivelEstoque.descricao.ilike(search_term),
            ConsumivelEstoque.categoria.ilike(search_term)
        ))
    
    consumiveis = query.order_by(ConsumivelEstoque.codigo_produto).all()

    # Carregar movimentações apenas para ADMIN
    movimentacoes = []
    if current_user.is_authenticated and current_user.role == 'admin':
        movimentacoes = MovimentacaoConsumivel.query.join(ConsumivelEstoque).order_by(MovimentacaoConsumivel.data_movimentacao.desc()).all()

    return render_template('consumivel.html', consumiveis=consumiveis, search_query=search_query, movimentacoes=movimentacoes)

@app.route('/consumivel/relatorio-movimentacoes')
@admin_only
@login_required
def relatorio_movimentacoes_consumivel_page():
    """Página do relatório de movimentações de consumíveis (apenas para admin)."""
    movimentacoes = MovimentacaoConsumivel.query.join(ConsumivelEstoque).order_by(MovimentacaoConsumivel.data_movimentacao.desc()).all()
    return jsonify({
        'movimentacoes': [
            {
                'id': mov.id,
                'tipo': mov.tipo,
                'quantidade': mov.quantidade,
                'data_movimentacao': mov.data_movimentacao.isoformat(),
                'observacao': mov.observacao,
                'usuario': mov.usuario,
                'setor_destino': mov.setor_destino,
                'codigo_produto': mov.consumivel.codigo_produto,
                'descricao': mov.consumivel.descricao,
                'unidade_medida': mov.consumivel.unidade_medida,
                'categoria': mov.consumivel.categoria
            }
            for mov in movimentacoes
        ]
    })

@app.route('/consumivel/importar', methods=['GET', 'POST'])
@admin_required
@login_required
def importar_consumivel():
    """Página e lógica para importar dados de consumíveis de uma planilha Excel."""
    if request.method == 'POST':
        if 'arquivo_excel' not in request.files:
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)
        
        file = request.files['arquivo_excel']
        if file.filename == '':
            flash('Nenhum arquivo selecionado!', 'danger')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                df = pd.read_excel(file)
                
                # Construir mapa de colunas normalizadas -> rótulo original (mais robusto)
                def normalize_text(s):
                    if not isinstance(s, str):
                        s = '' if pd.isna(s) else str(s)
                    s = s.strip()
                    # Remover símbolos ordinais e sinais que o Excel pode usar
                    s = s.replace('º', '').replace('°', '').replace('ª', '')
                    # Normalizar unicode (remover acentos)
                    s = unicodedata.normalize('NFKD', s)
                    s = ''.join(ch for ch in s if not unicodedata.combining(ch))
                    s = s.upper()
                    s = ' '.join(s.split())
                    return s

                orig_cols = [str(c) for c in df.columns]
                norm_map = {normalize_text(c): c for c in orig_cols}

                print(f"DEBUG: Colunas originais: {orig_cols}")
                print(f"DEBUG: Colunas normalizadas: {list(norm_map.keys())}")

                def find_col_by_patterns(*patterns):
                    for ncol, orig in norm_map.items():
                        for pat in patterns:
                            if pat in ncol:
                                return orig
                    return None

                # Procurar por colunas obrigatórias com variações (retorna rótulo original)
                col_n_produto = find_col_by_patterns('N PRODUTO', 'NPRODUTO', 'NUM PRODUTO', 'NUMERO PRODUTO')
                col_codigo = find_col_by_patterns('CODIGO PRODUTO', 'CODIGOPRODUTO', 'CODIGO')
                col_descricao = find_col_by_patterns('DESCRICAO DO PRODUTO', 'DESCRICAODOPRODUTO', 'DESCRICAO')
                col_unidade = find_col_by_patterns('UNIDADE MEDIDA', 'UNIDADE')

                # Verificar se todas as obrigatórias foram encontradas
                if not all([col_n_produto, col_codigo, col_descricao, col_unidade]):
                    flash(f'❌ Colunas obrigatórias faltando!\nColunas originais encontradas: {", ".join(orig_cols)}\nColunas normalizadas: {", ".join(list(norm_map.keys()))}', 'danger')
                    return redirect(request.url)
                
                sucesso = 0
                erro = 0
                
                for idx, row in df.iterrows():
                    try:
                        # Obter valores das colunas obrigatórias
                        n_produto = str(row.get(col_n_produto, '')).strip()
                        codigo = str(row.get(col_codigo, '')).strip()
                        desc = str(row.get(col_descricao, '')).strip()
                        unidade = str(row.get(col_unidade, '')).strip() or 'UN'
                        
                        # Validar obrigatórias
                        if not n_produto or not codigo or not desc:
                            erro += 1
                            continue
                        
                        # Identificar colunas opcionais usando nomes normalizados
                        col_status_estoque = find_col_by_patterns('STATUS ESTOQUE', 'STATUSESTOQUE', 'STATUS')
                        col_status_consumo = find_col_by_patterns('STATUS CONSUMO', 'STATUSCONSUMO')
                        col_categoria = find_col_by_patterns('CATEGORIA')
                        col_fornecedor = find_col_by_patterns('FORNECEDOR', 'FORNECEDOR 2', 'FORNECEDOR2')
                        col_fornecedor2 = find_col_by_patterns('FORNECEDOR 2', 'FORNECEDOR2')
                        col_valor = find_col_by_patterns('VALOR UNITARIO', 'VALORUNITARIO', 'VALOR')
                        col_lead = find_col_by_patterns('LEAD TIME', 'LEADTIME', 'TEMPO REPOSICAO', 'DIAS')
                        col_seg = find_col_by_patterns('ESTOQUE DE SEGURANCA', 'ESTOQUESEGURANCA', 'PERCENTUAL ESTOQUE')
                        col_minimo = find_col_by_patterns('ESTOQUE MINIMO', 'ESTOQUE MINIMO POR CAIXA', 'ESTOQUEMINIMO')
                        col_atual = find_col_by_patterns('ESTOQUE ATUAL', 'ESTOQUEATUAL', 'QUANTIDADE ATUAL')

                        # Obter valores das colunas opcionais usando rótulos originais (se encontrados)
                        status_estoque = str(row.get(col_status_estoque, 'Ativo')).strip() or 'Ativo'
                        status_consumo = str(row.get(col_status_consumo, 'Consumível')).strip() or 'Consumível'
                        categoria = str(row.get(col_categoria, '')).strip()
                        fornecedor = str(row.get(col_fornecedor, '')).strip()
                        fornecedor2 = str(row.get(col_fornecedor2, '')).strip()

                        try:
                            valor = float(row.get(col_valor, 0) or 0)
                        except:
                            valor = 0

                        try:
                            lead = int(row.get(col_lead, 7) or 7)
                        except:
                            lead = 7

                        try:
                            seg = float(row.get(col_seg, 0) or 0)
                        except:
                            seg = 0

                        try:
                            minimo = float(row.get(col_minimo, 0) or 0)
                        except:
                            minimo = 0

                        try:
                            atual = float(row.get(col_atual, 0) or 0)
                        except:
                            atual = 0
                        
                        # Criar ou atualizar
                        cons = ConsumivelEstoque.query.filter_by(codigo_produto=codigo).first()
                        if not cons:
                            cons = ConsumivelEstoque(
                                n_produto=n_produto,
                                codigo_produto=codigo,
                                descricao=desc,
                                unidade_medida=unidade,
                                status_estoque=status_estoque,
                                status_consumo=status_consumo,
                                categoria=categoria,
                                fornecedor=fornecedor,
                                fornecedor2=fornecedor2,
                                valor_unitario=valor,
                                lead_time=lead,
                                estoque_seguranca=seg,
                                estoque_minimo=minimo,
                                quantidade_atual=atual
                            )
                            db.session.add(cons)
                        else:
                            cons.n_produto = n_produto
                            cons.descricao = desc
                            cons.unidade_medida = unidade
                            cons.status_estoque = status_estoque
                            cons.status_consumo = status_consumo
                            cons.categoria = categoria
                            cons.fornecedor = fornecedor
                            cons.fornecedor2 = fornecedor2
                            cons.valor_unitario = valor
                            cons.lead_time = lead
                            cons.estoque_seguranca = seg
                            cons.estoque_minimo = minimo
                            cons.quantidade_atual = atual
                        
                        sucesso += 1
                    
                    except Exception as e:
                        erro += 1
                        print(f"Erro linha {idx + 2}: {e}")
                
                db.session.commit()
                flash(f'✅ {sucesso} consumível(is) importado(s)!', 'success')
                if erro > 0:
                    flash(f'⚠️ {erro} linha(s) ignorada(s).', 'warning')
                return redirect(url_for('consumivel'))
            
            except Exception as e:
                flash(f'❌ Erro: {str(e)}', 'danger')
                return redirect(request.url)
        else:
            flash('❌ Envie um arquivo .xlsx', 'danger')
            return redirect(request.url)
    
    return render_template('importar_consumivel.html')

@app.route('/consumivel/movimentacao', methods=['GET', 'POST'])
@admin_only
@login_required
def movimentacao_consumivel():
    """Página para registrar uma movimentação de consumível (entrada ou saída)."""
    if request.method == 'POST':
        consumivel_id = request.form.get('consumivel_id')
        tipo = request.form.get('tipo')
        quantidade = request.form.get('quantidade')
        setor_destino = request.form.get('setor_destino')
        observacao = request.form.get('observacao', '')

        if not consumivel_id or not tipo or not quantidade:
            flash('Consumível, tipo e quantidade são obrigatórios.', 'danger')
            return redirect(url_for('movimentacao_consumivel'))

        consumivel = ConsumivelEstoque.query.get_or_404(consumivel_id)

        try:
            quantidade = float(quantidade)
            if quantidade <= 0:
                flash('A quantidade deve ser um número positivo.', 'danger')
                return redirect(url_for('movimentacao_consumivel'))
        except (ValueError, TypeError):
            flash('A quantidade deve ser um número válido.', 'danger')
            return redirect(url_for('movimentacao_consumivel'))

        # --- LÓGICA DE ENTRADA ---
        if tipo == 'ENTRADA':
            consumivel.quantidade_atual += quantidade
            
            nova_movimentacao = MovimentacaoConsumivel(
                consumivel_id=consumivel.id,
                tipo='ENTRADA',
                quantidade=quantidade,
                usuario=current_user.username,
                setor_destino=setor_destino or 'Almoxarifado',
                observacao=observacao
            )
            db.session.add(nova_movimentacao)
            db.session.commit()
            flash('Entrada de consumível registrada com sucesso!', 'success')

        # --- LÓGICA DE SAÍDA ---
        elif tipo == 'SAIDA':
            if consumivel.quantidade_atual < quantidade:
                flash(f'Quantidade insuficiente. Disponível: {consumivel.quantidade_atual}', 'danger')
                return redirect(url_for('movimentacao_consumivel'))

            consumivel.quantidade_atual -= quantidade
            
            nova_movimentacao = MovimentacaoConsumivel(
                consumivel_id=consumivel.id,
                tipo='SAIDA',
                quantidade=quantidade,
                usuario=current_user.username,
                setor_destino=setor_destino,
                observacao=observacao
            )
            db.session.add(nova_movimentacao)
            db.session.commit()
            flash('Saída de consumível registrada com sucesso!', 'success')

        return redirect(url_for('movimentacao_consumivel'))

    # Para GET (exibir a página)
    consumiveis = ConsumivelEstoque.query.order_by(ConsumivelEstoque.descricao).all()
    return render_template('movimentacao_consumivel.html', consumiveis=consumiveis)

@app.route('/consumivel/editar/<int:consumivel_id>', methods=['GET', 'POST'])
@admin_required
@login_required
def editar_consumivel(consumivel_id):
    """Página para editar um consumível específico."""
    consumivel = ConsumivelEstoque.query.get_or_404(consumivel_id)

    if request.method == 'POST':
        try:
            consumivel.n_produto = request.form.get('n_produto', consumivel.n_produto)
            consumivel.status_estoque = request.form.get('status_estoque', consumivel.status_estoque)
            consumivel.status_consumo = request.form.get('status_consumo', consumivel.status_consumo)
            consumivel.descricao = request.form.get('descricao', consumivel.descricao)
            consumivel.unidade_medida = request.form.get('unidade_medida', consumivel.unidade_medida)
            consumivel.categoria = request.form.get('categoria', consumivel.categoria)
            consumivel.fornecedor = request.form.get('fornecedor', consumivel.fornecedor)
            consumivel.fornecedor2 = request.form.get('fornecedor2', consumivel.fornecedor2)
            consumivel.valor_unitario = float(request.form.get('valor_unitario', consumivel.valor_unitario))
            consumivel.lead_time = int(request.form.get('lead_time', consumivel.lead_time))
            consumivel.estoque_seguranca = float(request.form.get('estoque_seguranca', consumivel.estoque_seguranca))
            consumivel.estoque_minimo = float(request.form.get('estoque_minimo', consumivel.estoque_minimo))
            consumivel.quantidade_atual = float(request.form.get('quantidade_atual', consumivel.quantidade_atual))

            db.session.commit()
            flash('Consumível atualizado com sucesso!', 'success')
            return redirect(url_for('consumivel'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar o consumível: {e}', 'danger')

    return render_template('editar_consumivel.html', consumivel=consumivel)

@app.route('/consumivel/excluir/<int:consumivel_id>')
@admin_required
@login_required
def excluir_consumivel(consumivel_id):
    """Exclui um consumível do estoque."""
    consumivel = ConsumivelEstoque.query.get_or_404(consumivel_id)

    try:
        # Remove também as movimentações relacionadas
        MovimentacaoConsumivel.query.filter_by(consumivel_id=consumivel.id).delete()
        db.session.delete(consumivel)
        db.session.commit()
        flash(f'Consumível "{consumivel.descricao}" excluído com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir o consumível: {e}', 'danger')

    return redirect(url_for('consumivel'))

@app.route('/consumivel/historico/<int:consumivel_id>')
@admin_only
@login_required
def historico_consumivel(consumivel_id):
    """Exibe o histórico de movimentações de um consumível."""
    consumivel = ConsumivelEstoque.query.get_or_404(consumivel_id)
    movimentacoes = consumivel.movimentacoes.order_by(MovimentacaoConsumivel.data_movimentacao.desc()).all()
    return render_template('historico_consumivel.html', consumivel=consumivel, movimentacoes=movimentacoes)

@app.route('/api/consumivel/by-code/<string:codigo_produto>')
@login_required
def api_get_consumivel_by_code(codigo_produto):
    """Retorna os dados de um consumível pelo seu código em formato JSON."""
    consumivel = ConsumivelEstoque.query.filter(func.lower(ConsumivelEstoque.codigo_produto) == func.lower(codigo_produto)).first()
    
    if not consumivel:
        return jsonify({'error': 'Consumível não encontrado'}), 404
    
    return jsonify({
        'id': consumivel.id,
        'codigo_produto': consumivel.codigo_produto,
        'descricao': consumivel.descricao,
        'unidade_medida': consumivel.unidade_medida,
        'categoria': consumivel.categoria,
        'quantidade_atual': consumivel.quantidade_atual
    })


@app.route('/api/relatorio/movimentacoes-consumivel')
@admin_only
@login_required
def api_relatorio_movimentacoes_consumivel():
    """Retorna todas as movimentações de consumíveis em formato JSON."""
    movimentacoes = MovimentacaoConsumivel.query.join(ConsumivelEstoque).order_by(MovimentacaoConsumivel.data_movimentacao.desc()).all()
    
    dados = []
    for mov in movimentacoes:
        dados.append({
            'id': mov.id,
            'tipo': mov.tipo,
            'quantidade': mov.quantidade,
            'data_movimentacao': mov.data_movimentacao.isoformat(),
            'observacao': mov.observacao,
            'usuario': mov.usuario,
            'setor_destino': mov.setor_destino,
            'codigo_produto': mov.consumivel.codigo_produto,
            'descricao': mov.consumivel.descricao,
            'unidade_medida': mov.consumivel.unidade_medida,
            'categoria': mov.consumivel.categoria
        })
    
    return jsonify(dados)


@app.route('/consumivel/exportar')
@login_required
def exportar_consumivel():
    """Exporta todos os consumíveis para um arquivo Excel."""
    consumiveis = ConsumivelEstoque.query.order_by(ConsumivelEstoque.codigo_produto).all()
    
    # Preparar dados para o DataFrame na ordem correta para importação
    dados = []
    for consumivel in consumiveis:
        dados.append({
            'Nº PRODUTO': consumivel.n_produto,
            'STATUS ESTOQUE': consumivel.status_estoque,
            'STATUS CONSUMO': consumivel.status_consumo,
            'CÓDIGO PRODUTO': consumivel.codigo_produto,
            'DESCRIÇÃO DO PRODUTO': consumivel.descricao,
            'UNIDADE MEDIDA': consumivel.unidade_medida,
            'CATEGORIA': consumivel.categoria,
            'FORNECEDOR': consumivel.fornecedor,
            'FORNECEDOR 2': consumivel.fornecedor2,
            'VALOR UNITÁRIO': consumivel.valor_unitario,
            'LEAD TIME (DIAS ATRÁS)': consumivel.lead_time,
            '% ESTOQUE DE SEGURANÇA': consumivel.estoque_seguranca,
            'ESTOQUE MÍNIMO POR CAIXA': consumivel.estoque_minimo,
            'ESTOQUE ATUAL': consumivel.quantidade_atual,
        })
    
    # Criar DataFrame com as colunas na ordem correta
    df = pd.DataFrame(dados, columns=[
        'Nº PRODUTO',
        'STATUS ESTOQUE',
        'STATUS CONSUMO',
        'CÓDIGO PRODUTO',
        'DESCRIÇÃO DO PRODUTO',
        'UNIDADE MEDIDA',
        'CATEGORIA',
        'FORNECEDOR',
        'FORNECEDOR 2',
        'VALOR UNITÁRIO',
        'LEAD TIME (DIAS ATRÁS)',
        '% ESTOQUE DE SEGURANÇA',
        'ESTOQUE MÍNIMO POR CAIXA',
        'ESTOQUE ATUAL',
    ])
    
    # Criar arquivo Excel em memória
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Consumíveis', index=False)
        
        # Ajustar largura das colunas
        worksheet = writer.sheets['Consumíveis']
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    # Gerar resposta com o arquivo
    return make_response(
        output.getvalue(),
        200,
        {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="Consumiveis_{datetime.now().strftime("%d-%m-%Y_%H-%M-%S")}.xlsx"'
        }
    )

# --- ROTAS DE GERENCIAMENTO DE USUÁRIOS (ADMIN) ---

@app.route('/admin/usuarios')
@login_required
@admin_required
def gerenciar_usuarios():
    """Página para listar e gerenciar todos os usuários."""
    users = User.query.order_by(User.username).all()
    return render_template('admin_usuarios.html', users=users)

@app.route('/admin/usuario/novo', methods=['GET', 'POST'])
@login_required
@admin_required
def criar_usuario():
    """Página para criar um novo usuário."""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')

        if not username or not password or not role:
            flash('Todos os campos são obrigatórios.', 'danger')
            return redirect(url_for('criar_usuario'))

        if User.query.filter_by(username=username).first():
            flash('Este nome de usuário já está em uso.', 'danger')
            return redirect(url_for('criar_usuario'))

        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(username=username, password_hash=hashed_password, role=role)
        db.session.add(new_user)
        db.session.commit()
        flash(f'Usuário "{username}" criado com sucesso!', 'success')
        return redirect(url_for('gerenciar_usuarios'))

    return render_template('form_usuario.html', title="Criar Novo Usuário", action_url=url_for('criar_usuario'))

@app.route('/admin/usuario/editar/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(user_id):
    """Página para editar um usuário existente."""
    user = User.query.get_or_404(user_id)
    if request.method == 'POST':
        user.username = request.form.get('username')
        user.role = request.form.get('role')
        password = request.form.get('password')

        if password: # Só atualiza a senha se uma nova for fornecida
            user.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            flash('Senha atualizada com sucesso.', 'info')

        db.session.commit()
        flash(f'Usuário "{user.username}" atualizado com sucesso!', 'success')
        return redirect(url_for('gerenciar_usuarios'))

    return render_template('form_usuario.html', title="Editar Usuário", user=user, action_url=url_for('editar_usuario', user_id=user_id))

@app.route('/admin/usuario/excluir/<int:user_id>')
@login_required
@admin_required
def excluir_usuario(user_id):
    """Rota para excluir um usuário."""
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Você não pode excluir a si mesmo.', 'danger')
        return redirect(url_for('gerenciar_usuarios'))
    db.session.delete(user)
    db.session.commit()
    flash(f'Usuário "{user.username}" excluído com sucesso.', 'success')
    return redirect(url_for('gerenciar_usuarios'))

# --- ROTAS DE AUTENTICAÇÃO ---

@app.route('/login', methods=['GET', 'POST'])
def login():
    # Se o usuário já estiver logado, redireciona para o dashboard
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user and bcrypt.check_password_hash(user.password_hash, request.form.get('password')):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Login sem sucesso. Verifique o nome de usuário e a senha.', 'danger')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Você saiu do sistema com sucesso.', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Rota para registrar um novo usuário. Redireciona se já estiver logado."""
    # Se o usuário já estiver logado, não faz sentido registrar uma nova conta.
    if current_user.is_authenticated:
        flash('Você já está logado no sistema.', 'info')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if User.query.filter_by(username=username).first():
            flash('Este nome de usuário já existe. Por favor, escolha outro.', 'danger')
            return redirect(url_for('register'))

        hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
        new_user = User(username=username, password_hash=hashed_password) # 'admin' pode ser definido como padrão ou em outra lógica
        db.session.add(new_user)
        db.session.commit()
        flash('Sua conta foi criada com sucesso! Agora você pode fazer login.', 'success')
        return redirect(url_for('login'))
    return render_template('register.html')

# --- ROTA TEMPORÁRIA PARA PROMOVER UM USUÁRIO A ADMIN ---
# DEPOIS de usar, é RECOMENDADO remover ou comentar este código por segurança.
@app.route('/promote_to_admin')
@login_required
def promote_to_admin():
    """
    Rota temporária para promover o usuário logado a administrador.
    Acesse http://127.0.0.1:5000/promote_to_admin uma vez para se promover.
    """
    if current_user.role != 'admin':
        current_user.role = 'admin'
        db.session.commit()
        flash('Parabéns! Sua conta foi promovida para Administrador.', 'success')
    else:
        flash('Sua conta já é de um Administrador.', 'info')
    
    # Redireciona para a página de gerenciamento de usuários, que agora estará visível
    return redirect(url_for('gerenciar_usuarios'))
# --- FIM DA ROTA TEMPORÁRIA ---

# --- INICIALIZAÇÃO DA APLICAÇÃO ---
if __name__ == '__main__':
    # Cria as tabelas no banco de dados se não existirem
    with app.app_context():
        db.create_all()
        print("✅ Tabelas do banco de dados criadas/verificadas com sucesso!")
    
    app.run(debug=True)
       