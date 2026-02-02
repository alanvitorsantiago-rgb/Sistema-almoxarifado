import sqlite3
import openpyxl
from datetime import datetime
import sys

db_file = 'database.db'
estoque_file = 'relatorio_estoque_2026-01-07.xlsx'
consumiveis_file = 'Consumiveis_07-01-2026_16-50-34.xlsx'

print("🔄 SINCRONIZANDO BANCO COM PLANILHAS")
print("=" * 80)

try:
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    
    # ========== LIMPAR DADOS ANTIGOS ==========
    print("\n🗑️  DELETANDO DADOS ANTIGOS...")
    
    cursor.execute("DELETE FROM movimentacao_consumivel")
    cursor.execute("DELETE FROM movimentacao")
    cursor.execute("DELETE FROM estoque_detalhe")
    cursor.execute("DELETE FROM consumivel_estoque")
    cursor.execute("DELETE FROM item_estoque")
    
    print("  ✅ Todos os dados de itens, estoque e movimentações deletados")
    
    # ========== IMPORTAR ESTOQUE ==========
    print("\n📦 IMPORTANDO ESTOQUE DO ARQUIVO...")
    
    wb_estoque = openpyxl.load_workbook(estoque_file)
    ws_estoque = wb_estoque.active
    
    # Pular cabeçalho
    contador_itens = 0
    for row in ws_estoque.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Se a primeira coluna está vazia, para
            break
        
        try:
            # Mapear colunas da planilha de estoque
            codigo = str(row[0]).strip() if row[0] else ""
            descricao = str(row[1]).strip() if row[1] else ""
            endereco = str(row[2]).strip() if row[2] else None
            tipo = str(row[3]).strip() if row[3] else None
            un = str(row[4]).strip() if row[4] else None
            qtd_estoque = float(row[5]) if row[5] and str(row[5]).replace('.', '').replace('-', '').isdigit() else 0
            estoque_minimo = float(row[6]) if row[6] and str(row[6]).replace('.', '').replace('-', '').isdigit() else 5
            tempo_reposicao = int(row[7]) if row[7] and str(row[7]).isdigit() else 7
            
            if not codigo or not descricao:
                continue
            
            # Inserir item
            cursor.execute("""
                INSERT INTO item_estoque 
                (codigo, descricao, endereco, tipo, un, qtd_estoque, estoque_minimo, tempo_reposicao)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (codigo, descricao, endereco, tipo, un, qtd_estoque, estoque_minimo, tempo_reposicao))
            
            item_id = cursor.lastrowid
            
            # Se há quantidade, criar detalhe de estoque
            if qtd_estoque > 0:
                cursor.execute("""
                    INSERT INTO estoque_detalhe 
                    (item_estoque_id, lote, quantidade, estacao)
                    VALUES (?, ?, ?, ?)
                """, (item_id, 'LOTE-IMPORTAÇÃO', qtd_estoque, 'Almoxarifado'))
            
            contador_itens += 1
            print(f"  ✅ {codigo}: {descricao} ({qtd_estoque} {un})")
        
        except Exception as e:
            print(f"  ⚠️  Erro ao processar linha: {e}")
            continue
    
    print(f"\n✅ Total de itens importados: {contador_itens}")
    
    # ========== IMPORTAR CONSUMÍVEIS ==========
    print("\n🧴 IMPORTANDO CONSUMÍVEIS DO ARQUIVO...")
    
    wb_consumiveis = openpyxl.load_workbook(consumiveis_file)
    ws_consumiveis = wb_consumiveis.active
    
    contador_consumiveis = 0
    for row in ws_consumiveis.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # Se a primeira coluna está vazia, para
            break
        
        try:
            # Mapear colunas da planilha de consumíveis
            n_produto = str(row[0]).strip() if row[0] else ""
            status_estoque = str(row[1]).strip() if row[1] else None
            status_consumo = str(row[2]).strip() if row[2] else None
            codigo_produto = str(row[3]).strip() if row[3] else ""
            descricao = str(row[4]).strip() if row[4] else ""
            unidade_medida = str(row[5]).strip() if row[5] else None
            categoria = str(row[6]).strip() if row[6] else None
            fornecedor = str(row[7]).strip() if row[7] else None
            fornecedor2 = str(row[8]).strip() if row[8] else None
            valor_unitario = float(row[9]) if row[9] and str(row[9]).replace('.', '').replace('-', '').isdigit() else 0
            lead_time = int(row[10]) if row[10] and str(row[10]).isdigit() else 7
            estoque_seguranca = float(row[11]) if row[11] and str(row[11]).replace('.', '').replace('-', '').isdigit() else 0
            estoque_minimo = float(row[12]) if row[12] and str(row[12]).replace('.', '').replace('-', '').isdigit() else 0
            quantidade_atual = float(row[13]) if row[13] and str(row[13]).replace('.', '').replace('-', '').isdigit() else 0
            
            if not n_produto or not codigo_produto or not descricao:
                continue
            
            # Inserir consumível
            cursor.execute("""
                INSERT INTO consumivel_estoque 
                (n_produto, status_estoque, status_consumo, codigo_produto, descricao, 
                 unidade_medida, categoria, fornecedor, fornecedor2, valor_unitario,
                 lead_time, estoque_seguranca, estoque_minimo, quantidade_atual)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (n_produto, status_estoque, status_consumo, codigo_produto, descricao,
                 unidade_medida, categoria, fornecedor, fornecedor2, valor_unitario,
                 lead_time, estoque_seguranca, estoque_minimo, quantidade_atual))
            
            contador_consumiveis += 1
            print(f"  ✅ {n_produto}: {descricao} ({quantidade_atual} {unidade_medida})")
        
        except Exception as e:
            print(f"  ⚠️  Erro ao processar linha: {e}")
            continue
    
    print(f"\n✅ Total de consumíveis importados: {contador_consumiveis}")
    
    # ========== COMMIT ==========
    conn.commit()
    
    # ========== VERIFICAÇÃO FINAL ==========
    print("\n" + "=" * 80)
    print("📊 RESUMO FINAL:")
    
    cursor.execute("SELECT COUNT(*) FROM item_estoque")
    total_itens = cursor.fetchone()[0]
    
    cursor.execute("SELECT SUM(qtd_estoque) FROM item_estoque")
    total_qtd = cursor.fetchone()[0] or 0
    
    cursor.execute("SELECT COUNT(*) FROM consumivel_estoque")
    total_consumiveis = cursor.fetchone()[0]
    
    print(f"  📦 Itens de Estoque: {total_itens}")
    print(f"  📊 Quantidade Total: {total_qtd}")
    print(f"  🧴 Consumíveis: {total_consumiveis}")
    
    print("\n✅ SINCRONIZAÇÃO CONCLUÍDA COM SUCESSO!")
    print("=" * 80)
    
    conn.close()

except Exception as e:
    print(f"❌ ERRO GERAL: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)
