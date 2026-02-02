#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Teste de exportação Excel - simular o que será gerado."""

from app import db, app
from models import ItemEstoque, EstoqueDetalhe
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

with app.app_context():
    # Buscar alguns itens para testar
    query = db.session.query(EstoqueDetalhe).join(ItemEstoque).order_by(
        ItemEstoque.codigo, 
        EstoqueDetalhe.data_entrada
    ).limit(5).all()
    
    print("📊 SIMULAÇÃO DE EXPORTAÇÃO EXCEL")
    print("=" * 120)
    print()
    
    # Definição de colunas
    colunas_ordem = [
        'CÓDIGO', 
        'CÓDIGO OPCIONAL', 
        'TIPO', 
        'DESCRIÇÃO', 
        'LOCAL', 
        'UN.', 
        'DIMENSÃO', 
        'CLIENTE', 
        'LOTE', 
        'ITEM NF', 
        'NF', 
        'VALIDADE', 
        'ESTAÇÃO', 
        'QTD ESTOQUE', 
        'DATA ENTRADA'
    ]
    
    # Configuração do Excel (Correção: Criar o arquivo real)
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"
    
    # Cabeçalhos
    print("CABEÇALHOS (Row 1):")
    for i, col in enumerate(colunas_ordem, 1):
        print(f"  Col {i:2d}: {col}")
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    print()
    
    # Dados
    print("DADOS (Rows 2+):")
    print()
    
    for num_linha, detalhe in enumerate(query, 2):
        item = detalhe.item_estoque
        
        dados_linha = {
            'CÓDIGO': str(item.codigo).strip() if item.codigo else '',
            'CÓDIGO OPCIONAL': str(item.codigo_opcional).strip() if item.codigo_opcional else '',
            'TIPO': str(item.tipo).strip() if item.tipo else '',
            'DESCRIÇÃO': str(item.descricao).strip() if item.descricao else '',
            'LOCAL': str(item.endereco).strip() if item.endereco else '',
            'UN.': str(item.un).strip() if item.un else 'UN',
            'DIMENSÃO': str(item.dimensao).strip() if item.dimensao else '',
            'CLIENTE': str(item.cliente).strip() if item.cliente else '',
            'LOTE': str(detalhe.lote).strip() if detalhe.lote else '',
            'ITEM NF': str(detalhe.item_nf).strip() if detalhe.item_nf else '',
            'NF': str(detalhe.nf).strip() if detalhe.nf else '',
            'VALIDADE': detalhe.validade.strftime('%d/%m/%Y') if detalhe.validade else '',
            'ESTAÇÃO': str(detalhe.estacao).strip() if detalhe.estacao else '',
            'QTD ESTOQUE': round(float(detalhe.quantidade), 2) if detalhe.quantidade else 0,
            'DATA ENTRADA': detalhe.data_entrada.strftime('%d/%m/%Y %H:%M:%S') if detalhe.data_entrada else '',
        }
        
        print(f"Row {num_linha}:")
        for num_col, nome_col in enumerate(colunas_ordem, 1):
            valor = dados_linha.get(nome_col, '')
            
            # Escrever no Excel (Correção: Garantir que o valor vai para a coluna certa)
            ws.cell(row=num_linha, column=num_col, value=valor)
            
            # Truncar valores longos para visualização
            valor_display = str(valor)[:50] + ('...' if len(str(valor)) > 50 else '')
            print(f"  {nome_col:20s} → {valor_display}")
        print()
    
    # Ajustar largura das colunas automaticamente
    for i, col in enumerate(colunas_ordem, 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    # Salvar o arquivo para conferência
    arquivo_saida = "teste_exportacao_resultado.xlsx"
    wb.save(arquivo_saida)
    
    print("=" * 120)
    print(f"✅ EXPORTAÇÃO CONCLUÍDA! Arquivo gerado: {arquivo_saida}")
    print("✅ Abra este arquivo no Excel para confirmar se os títulos e dados estão corretos.")
